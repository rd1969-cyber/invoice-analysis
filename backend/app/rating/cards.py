"""Generic loader for InXpress carrier rate-sheet .xls files.

All the cards (DHL, Canpar, Purolator) share one block layout per sheet:

    Prepared for: INXPRESS
    <Product name>                      e.g. "Purolator Express", "Express Parcel Single"
    Value in CAD
    Weight(lb)  <zone> <zone> ...       zone labels: N1-N14 / D01-D16 / 1-16
    1.0   <price> <price> ...           weight breakpoint rows
    ...
    Non-Document above N lb (Multiply shipment weight by zone rate)
    Weight(lb)  <zone> ...
    N.x   <per-unit rate> ...           overage: price per unit above the table

This one loader handles all of them; carrier-specific behaviour (DIM factors,
which product to use per scope, zone resolution) lives in ``carriers.py``.

Prices are stored as integer cents. The .xls files are OLE2 but trip a known
xlrd false-positive, hence ``ignore_workbook_corruption=True``.
"""
from __future__ import annotations

import os
from dataclasses import dataclass, field
from decimal import Decimal

import xlrd

_SKIP_PREFIXES = ("prepared", "value in", "all rates", "rates for specific")


def _cents(v) -> int | None:
    if v is None or v == "":
        return None
    if isinstance(v, str):
        v = v.replace(",", "").strip()
        if not v:
            return None
    try:
        return int((Decimal(str(v)) * 100).to_integral_value())
    except Exception:
        return None


def _is_num(v) -> bool:
    if isinstance(v, (int, float)):
        return True
    if isinstance(v, str):
        try:
            float(v.replace(",", ""))
            return True
        except ValueError:
            return False
    return False


def _zone_label(v) -> str:
    """Normalize a zone header cell: 1.0 -> '1', 'D01' -> 'D01', 'N1' -> 'N1'."""
    if isinstance(v, float) and v.is_integer():
        return str(int(v))
    if isinstance(v, (int,)):
        return str(v)
    return str(v).strip()


@dataclass
class Product:
    name: str
    unit: str = "lb"  # "lb" or "kg"
    zones: list[str] = field(default_factory=list)
    breakpoints: dict[str, list[tuple[float, int]]] = field(default_factory=dict)
    overage: dict[str, int] = field(default_factory=dict)  # per-unit rate above table

    def quote_base_cents(self, zone: str, weight: float) -> tuple[int | None, str]:
        bp = self.breakpoints.get(zone)
        if not bp:
            return None, f"zone {zone} not in {self.name}"
        for max_w, price in bp:
            if weight <= max_w:
                return price, f"{weight:g}{self.unit} <= {max_w:g} band @ {self.name}/{zone}"
        ov = self.overage.get(zone)
        if ov is not None:
            return int(round(ov * weight)), f"{weight:g}{self.unit} x {ov/100:.2f}/{self.unit} overage"
        return None, f"{weight:g}{self.unit} exceeds table for {self.name}/{zone}"


def _norm(s: str) -> str:
    """Lowercase, keep only a-z0-9 — so 'Purolator Ground®' == 'Purolator Ground'."""
    return "".join(ch for ch in s.lower() if ch.isalnum())


@dataclass
class RateCardData:
    carrier: str = ""
    currency: str = "CAD"
    products: dict[str, Product] = field(default_factory=dict)

    def get(self, name: str) -> Product | None:
        """Find a product tolerant of trademark symbols / spacing / case."""
        if name in self.products:
            return self.products[name]
        want = _norm(name)
        for k, p in self.products.items():
            nk = _norm(k)
            if nk == want or nk.startswith(want):
                return p
        return None


def adjust_card(card: RateCardData, factor_pct: float) -> RateCardData:
    """Return a copy of the card with all base rates scaled by (1 + factor_pct).

    Lets the user nudge a carrier's base rates up/down in the app (e.g. +5% or
    a negotiated -8%) without re-uploading a file. factor_pct is a fraction,
    e.g. 0.05 for +5%.
    """
    if not factor_pct:
        return card
    mult = 1 + factor_pct
    out = RateCardData(carrier=card.carrier, currency=card.currency)
    for name, p in card.products.items():
        np = Product(name=p.name, unit=p.unit, zones=list(p.zones))
        np.breakpoints = {
            z: [(w, int(round(c * mult))) for w, c in bps]
            for z, bps in p.breakpoints.items()
        }
        np.overage = {z: int(round(c * mult)) for z, c in p.overage.items()}
        out.products[name] = np
    return out


Grid = list[list]  # rows of cells (str / float / None)


def parse_grid(rows: Grid, carrier: str = "", currency: str = "CAD") -> RateCardData:
    """Parse a normalized grid (from Excel or PDF) into a RateCardData.

    Robust to both single-cell layouts (Excel/PDF tables, where a product name is
    one cell) and whitespace-split layouts (PDF text, where 'DHL Express - Package'
    arrives as several cells): product names are re-joined from the row, and the
    overage / unit triggers scan the whole row's text.
    """
    card = RateCardData(carrier=carrier, currency=currency)
    current: Product | None = None
    expect_product = False
    is_overage = False

    for row in rows:
        nonempty = [str(c).strip() for c in row if c is not None and str(c).strip() != ""]
        if not nonempty:
            continue
        c0 = str(row[0]).strip() if row[0] is not None and str(row[0]).strip() != "" else nonempty[0]
        low0 = c0.lower()
        rowtext = " ".join(nonempty).lower()

        if low0.startswith("prepared"):
            expect_product = True
            continue
        if expect_product:
            if not low0.startswith("value"):
                name = " ".join(nonempty)
                current = Product(name=name)
                card.products[name] = current
                is_overage = False
                expect_product = False
            continue
        if low0.startswith("value"):
            continue
        if low0.startswith("weight("):
            zones = [_zone_label(c) for c in row[1:] if c is not None and str(c).strip() != ""]
            if current is not None:
                current.unit = "kg" if "kg" in low0 else "lb"
                if not current.zones:
                    current.zones = zones
            continue
        if "above" in rowtext and ("lb" in rowtext or "kg" in rowtext):
            is_overage = True
            continue
        if current is not None and _is_num(row[0]):
            weight = float(str(row[0]).replace(",", ""))
            for i, zone in enumerate(current.zones):
                price = _cents(row[i + 1]) if i + 1 < len(row) else None
                if price is None:
                    continue
                if is_overage:
                    current.overage[zone] = price
                else:
                    current.breakpoints.setdefault(zone, []).append((weight, price))
            continue

    for prod in card.products.values():
        for zone in prod.breakpoints:
            prod.breakpoints[zone].sort()
    return card


# --------------------------------------------------------------------------- #
# Format readers -> {sheet_name: Grid}
# --------------------------------------------------------------------------- #
def _read_xls(path: str) -> dict[str, Grid]:
    bk = xlrd.open_workbook(path, ignore_workbook_corruption=True)
    out: dict[str, Grid] = {}
    for name in bk.sheet_names():
        sh = bk.sheet_by_name(name)
        out[name] = [[sh.cell_value(r, c) for c in range(sh.ncols)] for r in range(sh.nrows)]
    return out


def _read_xlsx(path: str) -> dict[str, Grid]:
    from openpyxl import load_workbook

    wb = load_workbook(path, read_only=True, data_only=True)
    out: dict[str, Grid] = {}
    for ws in wb.worksheets:
        out[ws.title] = [list(r) for r in ws.iter_rows(values_only=True)]
    wb.close()
    return out


def _read_pdf(path: str) -> dict[str, Grid]:
    import pdfplumber

    rows: Grid = []
    with pdfplumber.open(path) as pdf:
        for page in pdf.pages:
            tables = page.extract_tables()
            if tables:
                for tbl in tables:
                    for r in tbl:
                        rows.append([(c if c is not None else "") for c in r])
            else:  # text fallback: split each line on whitespace
                for line in (page.extract_text() or "").splitlines():
                    if line.strip():
                        rows.append(line.split())
    return {"PDF": rows}


def _pick_sheet(grids: dict[str, Grid], sheet: str | None) -> Grid:
    if sheet and sheet in grids:
        return grids[sheet]
    # Prefer the outbound parcel sheet, then the single-sheet domestic layout.
    for pref in ("OUTBOUND", "DIFFERENT", "PDF"):
        if pref in grids:
            return grids[pref]
    return next(iter(grids.values())) if grids else []


def load_any(path: str, carrier: str = "", sheet: str | None = None) -> RateCardData:
    """Load a rate card from .xls, .xlsx, or .pdf, auto-selecting the sheet."""
    ext = os.path.splitext(path)[1].lower()
    if ext == ".xls":
        grids = _read_xls(path)
    elif ext in (".xlsx", ".xlsm"):
        grids = _read_xlsx(path)
    elif ext == ".pdf":
        grids = _read_pdf(path)
    else:
        raise ValueError(f"Unsupported rate-card format: {ext} (use .xls, .xlsx, or .pdf)")
    return parse_grid(_pick_sheet(grids, sheet), carrier=carrier)


def load_card(path: str, sheet: str, carrier: str = "") -> RateCardData:
    """Backwards-compatible .xls loader (now delegates to the grid parser)."""
    return parse_grid(_pick_sheet(_read_xls(path), sheet), carrier=carrier)
