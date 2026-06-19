"""Branded Excel export of the rate comparison.

Two modes (per the spec):
  * internal  — full picture: my carrier, my cost, difference, margin, markup
  * customer  — savings story only: their current price, your price, savings.
                NO cost or margin shown.

Difference uses the requested colour convention: RED when my cost is HIGH
(uncompetitive), black/dark when competitive. Styled with InXpress brand colours.
"""
from __future__ import annotations

import io

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from app import brand

_BLUE = brand.MIDNIGHT_BLUE.lstrip("#")
_GREEN = brand.SPRING_GREEN.lstrip("#")
_RED = brand.RED.lstrip("#")
_SURFACE = brand.SURFACE_BLUE.lstrip("#")

_HEADER_FILL = PatternFill("solid", fgColor=_BLUE)
_SUBHEAD_FILL = PatternFill("solid", fgColor=_SURFACE)
_WHITE_BOLD = Font(name="Calibri", bold=True, color="FFFFFF", size=12)
_BLUE_BOLD = Font(name="Calibri", bold=True, color=_BLUE)
_RED_FONT = Font(name="Calibri", color=_RED, bold=True)
_DARK_FONT = Font(name="Calibri", color=_BLUE)
_MONEY = "#,##0.00"
_PCT = "0%"


def _title_block(ws, title: str, mode: str, settings: dict) -> int:
    # Blue logo on white (brand: blue logo on light backgrounds), above the band.
    logo = brand.logo_path(on_dark=False)
    base = 1
    if logo:
        try:
            from openpyxl.drawing.image import Image as XLImage

            img = XLImage(logo)
            img.width, img.height = 200, 41  # keep 1000x206 aspect
            ws.add_image(img, "A1")
            ws.row_dimensions[1].height = 34
            base = 2
        except Exception:
            base = 1

    ws.merge_cells(f"A{base}:F{base}")
    ws[f"A{base}"] = f"{brand.APP_NAME} — {title}"
    ws[f"A{base}"].font = Font(name="Calibri", bold=True, color="FFFFFF", size=16)
    ws[f"A{base}"].fill = _HEADER_FILL
    ws[f"A{base}"].alignment = Alignment(vertical="center", indent=1)
    ws.row_dimensions[base].height = 30

    c = brand.CONTACT
    ws.merge_cells(f"A{base+1}:F{base+1}")
    ws[f"A{base+1}"] = f"{c['business_name']}  |  {c['email']}  |  {c['phone']}"
    ws[f"A{base+1}"].font = Font(name="Calibri", color="FFFFFF", size=9)
    ws[f"A{base+1}"].fill = PatternFill("solid", fgColor=_GREEN)
    ws.row_dimensions[base + 1].height = 16

    row = base + 3
    bits = []
    if "target_customer_savings" in settings:
        bits.append(f"Target customer savings: {settings['target_customer_savings']:.0%}")
    if mode == "internal" and "min_margin_pct" in settings:
        bits.append(f"Min margin floor: {settings['min_margin_pct']:.0%}")
    if bits:
        ws.cell(row, 1, "  •  ".join(bits)).font = Font(name="Calibri", italic=True, color="666666")
        row += 2
    return row


def _kpis(ws, start_row: int, summary: dict, mode: str) -> int:
    r = start_row
    ws.cell(r, 1, "Summary").font = _BLUE_BOLD
    r += 1
    if mode == "internal":
        pairs = [
            ("Shipments", summary["shipments"]),
            ("Serviceable", summary["serviceable"]),
            ("Winnable lanes (beat model)", summary["winnable"]),
            ("Competitor spend", f"${summary['competitor_total']:,.2f}"),
            ("My cost", f"${summary['my_cost_total']:,.2f}"),
            ("Margin — beat model", f"${summary['total_margin']:,.2f}"),
            ("Margin — margin model", f"${summary.get('margin_total_margin', 0):,.2f}"),
        ]
    else:
        pairs = [
            ("Shipments reviewed", summary["shipments"]),
            ("Lanes we can save you on", summary["winnable"]),
            ("Total estimated savings", f"${summary['total_customer_savings']:,.2f}"),
        ]
    for label, val in pairs:
        ws.cell(r, 1, label).font = Font(name="Calibri", color="666666")
        ws.cell(r, 2, val).font = _BLUE_BOLD
        r += 1
    return r + 1


def _autofit(ws, widths: dict[int, int]):
    for col, w in widths.items():
        ws.column_dimensions[get_column_letter(col)].width = w


def build_workbook(records: list[dict], summary: dict, mode: str, settings: dict) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Comparison" if mode == "internal" else "Your savings"
    ws.sheet_view.showGridLines = False

    title = "Internal margin report" if mode == "internal" else "Your savings summary"
    row = _title_block(ws, title, mode, settings)
    row = _kpis(ws, row, summary, mode)

    basis = settings.get("pricing_basis", "beat")  # 'beat' or 'costplus'
    if mode == "internal":
        # Full picture: every carrier's cost side by side + BOTH pricing models.
        headers = ["Tracking", "Pickup", "Delivery", "Actual wt", "Billable wt", "Dims",
                   "Wt basis", "Competitor svc", "Scope", "They pay",
                   "Purolator", "Canpar", "DHL", "Best", "My cost", "Difference", "Status",
                   "Beat sell", "Beat margin", "Beat %", "Margin sell", "Margin $", "Margin %"]
        keys = ["tracking", "pickup", "delivery", "actual_wt", "billable_wt", "dims",
                "weight_basis", "competitor_service", "scope", "competitor_pays",
                "Purolator_cost", "Canpar_cost", "DHL_cost", "my_carrier", "my_cost",
                "difference", "status", "beat_sell", "beat_margin", "beat_margin_pct",
                "mgn_sell", "mgn_margin", "mgn_margin_pct"]
        money_cols = {10, 11, 12, 13, 15, 16, 18, 19, 21, 22}
        pct_cols = {20, 23}
        data = records
    else:
        headers = ["Tracking", "Service", "Current price", "Your InXpress price",
                   "You save", "% saved"]
        keys = ["tracking", "competitor_service", "competitor_pays", "_your_price",
                "_you_save", "_pct_saved"]
        money_cols = {3, 4, 5}
        pct_cols = {6}
        sell_key = "mgn_sell" if basis == "margin" else "beat_sell"
        save_key = "mgn_savings" if basis == "margin" else "beat_savings"
        # customer view: only lanes where we can offer a saving (price <= their price)
        data = [r for r in records if r.get(save_key) is not None and r[save_key] >= 0
                and r.get(sell_key)]
        for r in data:
            r["_your_price"] = r[sell_key]
            r["_you_save"] = r[save_key]
            r["_pct_saved"] = (r[save_key] / r["competitor_pays"]) if r["competitor_pays"] else 0

    head_row = row
    for c, h in enumerate(headers, 1):
        cell = ws.cell(head_row, c, h)
        cell.font = _WHITE_BOLD
        cell.fill = _HEADER_FILL
        cell.alignment = Alignment(vertical="center")
    ws.row_dimensions[head_row].height = 20
    row += 1

    for rec in data:
        is_high = rec.get("status") == "HIGH"
        for c, key in enumerate(keys, 1):
            val = rec.get(key)
            cell = ws.cell(row, c, val)
            if c in money_cols and isinstance(val, (int, float)):
                cell.number_format = _MONEY
            if c in pct_cols and isinstance(val, (int, float)):
                cell.number_format = _PCT
            # red/black convention on difference + status (internal)
            if mode == "internal" and key in ("difference", "status"):
                cell.font = _RED_FONT if is_high else _DARK_FONT
            elif mode == "customer" and key in ("_you_save", "_pct_saved"):
                cell.font = Font(name="Calibri", bold=True, color=_GREEN.upper())
        row += 1

    widths = ({1: 19, 2: 8, 3: 8, 4: 8, 5: 9, 6: 10, 7: 9, 8: 14, 9: 13, 10: 9, 11: 9,
               12: 9, 13: 9, 14: 9, 15: 9, 16: 9, 17: 7, 18: 9, 19: 9, 20: 7, 21: 9, 22: 9,
               23: 7}
              if mode == "internal" else {1: 20, 2: 18, 3: 13, 4: 17, 5: 11, 6: 9})
    _autofit(ws, widths)
    ws.freeze_panes = ws.cell(head_row + 1, 1)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
